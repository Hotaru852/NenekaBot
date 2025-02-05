import requests
import re
from openpyxl import load_workbook
import pandas as pd
import io
import json
import hashlib

SHEET_ID = '1JjK7Ws4gfzKChRs5ueoxEZVN5SXK10nhDC1-nbm0NUs'
GID = '303232073'

url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&id={SHEET_ID}&gid={GID}"

def fetch_data():
    response = requests.get(url)
    if response.status_code == 200:
        return response.content
    else:
        return None

def preprocess_data(data):
    if not data:
        data = fetch_data()

    # Read the data into a pandas dataframe
    df = pd.read_csv(io.StringIO(data.decode('utf-8')))

    # Write the dataframe to an excel file in memory
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Load the in-memory excel file into openpyxl
    excel_buffer.seek(0)  # Rewind the buffer
    work_book = load_workbook(filename=io.BytesIO(excel_buffer.getvalue()))

    sheet = work_book.active

    rows_to_delete = []

    for row in range(2, sheet.max_row + 1):
        cell = sheet[f"C{row}"]

        if cell.value:
            cell.value = ''.join(cell.value.split(','))

        cell = sheet[f"A{row}"]

        if cell.value and cell.value != ' ':
            rows_to_delete.append(row - 1)

    count = 0
    for row in rows_to_delete:
        sheet.delete_rows(row - 2 * count, 2)
        count += 1
        
    unit_template = {
        'Name': None,
        'Nickname': None,
        'Rarity': None,
        'Element': None,
        'Position': None,
        'Role': None,
        'Initial Movement': None,
        'Loop Pattern': None,
        'UB Initial Movement': None,
        'UB Loop Pattern': None,
        'Union Burst': None,
        'Union Burst Description': None,
        'Union Burst Damage Distribution': None,
        'Union Burst+': None,
        'Union Burst+ Description': None,
        'Union Burst+ Damage Distribution': None,
        'Skill 1': None,
        'Skill 1 Description': None,
        'Skill 1 Damage Distribution': None,
        'Skill 1+': None,
        'Skill 1+ Description': None,
        'Skill 1 Damage+ Distribution': None,
        'Skill 2': None,
        'Skill 2 Description': None,
        'Skill 2 Damage Distribution': None,
        'Skill 2+': None,
        'Skill 2+ Description': None,
        'Skill 2+ Damage Distribution': None,
        'SP 1': None,
        'SP 1 Description': None,
        'SP 1 Damage Distribution': None, 
        'SP 1+': None,
        'SP 1+ Description': None,
        'SP 1+ Damage Distribution': None, 
        'SP 2': None,
        'SP 2 Description': None,
        'SP 2 Damage Distribution': None, 
        'SP 2+': None,
        'SP 2+ Description': None,
        'SP 2+ Damage Distribution': None, 
        'SP 3': None,
        'SP 3 Description': None,
        'SP 3 Damage Distribution': None,
        'SP Skill': None,
        'SP Skill Description': None,
        'EX Skill': None,
        'EX Skill Description': None,
        'Unique Equipment 1': None,
        'Unique Equipment 1 Stats': None,
        'Unique Equipment 2': None,
        'Unique Equipment 2 Stats': None,
        'Misc. Information': None,
        'Review': None
    }

    units_data = []

    current_unit_index = -1
    current_skill = None
    current_section = None
    has_damage_distribution = False

    for row in sheet.iter_rows(values_only=True):
        row = ','.join(str(cell) if cell is not None else '' for cell in row)
        row = re.sub(r',+', ',', row.strip(',').replace('"', ''))
        strings = row.split(',')

        if (re.match(r".+Level \d+,★+,Rank \d+.+", row)):
            units_data.append(unit_template.copy())
            current_unit_index += 1
            current_section = None
            units_data[current_unit_index]['Name'] = strings[0]
            units_data[current_unit_index]['Rarity'] = strings[2]
            units_data[current_unit_index]['Element'] = strings[4]
            if '&' not in strings[0]:
                if len(strings[0].split()) > 1:
                    abbreviation = None
                    parts = strings[0].split()
                    if len(parts) == 2:
                        if parts[1] == '(Christmas)':
                            abbreviation = 'X'
                        elif parts[1] == '(Sarasaria)':
                            abbreviation = 'SA'
                        elif parts[1] == '(Commander)':
                            abbreviation = 'CO'
                        elif parts[1] == '(Spring)':
                            abbreviation = 'SP'
                        else:
                            abbreviation = parts[1][1]
                    else:
                        if parts[1] + parts[2] == '(RitualGarment)':
                            abbreviation = 'C'
                        else:
                            abbreviation = parts[1][1] + parts[2][0]
                    units_data[current_unit_index]['Nickname'] = abbreviation + parts[0]
                else:
                    units_data[current_unit_index]['Nickname'] = strings[0]
            else:
                if strings[0] == 'Misogi & Mimi & Kyouka':
                    units_data[current_unit_index]['Nickname'] = 'LLtrio'
                else:
                    units_data[current_unit_index]['Nickname'] = ''.join(strings[0].replace('&', '').split())
        else:
            if row.startswith('Position') and not current_section:
                units_data[current_unit_index]['Position'] = strings[1]
            elif row.startswith('Role') and not current_section:
                units_data[current_unit_index]['Role'] = ','.join(strings[1:])
            elif row.startswith('Initial Movement') and not current_section:
                units_data[current_unit_index]['Initial Movement'] = [string.replace('Hidden Skill', 'SP Skill') for string in strings[1:]]
            elif ('Initial' in row or 'Movement' in row) and not current_section and not current_skill:
                units_data[current_unit_index]['UB Initial Movement'] = strings[1:]
            elif row.startswith('Loop Pattern') and not current_section:
                units_data[current_unit_index]['Loop Pattern'] = strings[1:]
            elif ('Loop' in row or 'Pattern' in row) and not current_section and not current_skill:
                units_data[current_unit_index]['UB Loop Pattern'] = strings[1:]
            elif row.startswith('Damage Distribution') and not current_section:
                units_data[current_unit_index][current_skill + ' Damage Distribution'] = strings[0] + ':'
                units_data[current_unit_index][current_skill + ' Damage Distribution'] += ' ' + ','.join(strings[1:])
                has_damage_distribution = True
            elif row.startswith('Union Burst') and not current_section:
                units_data[current_unit_index]['Union Burst'] = strings[1]
                units_data[current_unit_index]['Union Burst Description'] = ','.join(strings[2:])
                current_skill = "Union Burst"
                has_damage_distribution = False
            elif row.startswith("★6 Union Burst") and not current_section:
                units_data[current_unit_index]['Union Burst+'] = strings[1]
                units_data[current_unit_index]['Union Burst+ Description'] = ','.join(strings[2:])
                current_skill = "Union Burst+"
                has_damage_distribution = False
            elif row.startswith("Skill 1+") and not current_section:
                units_data[current_unit_index]['Skill 1+'] = strings[1]
                units_data[current_unit_index]['Skill 1+ Description'] = ','.join(strings[2:])
                current_skill = "Skill 1+"
                has_damage_distribution = False
            elif row.startswith("Skill 1") and not current_section:
                units_data[current_unit_index]['Skill 1'] = strings[1]
                units_data[current_unit_index]['Skill 1 Description'] = ','.join(strings[2:])
                current_skill = "Skill 1"
                has_damage_distribution = False
            elif row.startswith("Skill 2+") and not current_section:
                units_data[current_unit_index]['Skill 2+'] = strings[1]
                units_data[current_unit_index]['Skill 2+ Description'] = ','.join(strings[2:])
                current_skill = "Skill 2+"
                has_damage_distribution = False
            elif row.startswith("Skill 2") and not current_section:
                units_data[current_unit_index]['Skill 2'] = strings[1]
                units_data[current_unit_index]['Skill 2 Description'] = ','.join(strings[2:])
                current_skill = "Skill 2"
                has_damage_distribution = False
            elif row.startswith("SP 1+") and not current_section:
                units_data[current_unit_index]['SP 1+'] = strings[1]
                units_data[current_unit_index]['SP 1+ Description'] = ','.join(strings[2:])
                current_skill = "SP 1+"
                has_damage_distribution = False
            elif row.startswith("SP 1") and not current_section:
                units_data[current_unit_index]['SP 1'] = strings[1]
                units_data[current_unit_index]['SP 1 Description'] = ','.join(strings[2:])
                current_skill = "SP 1"
                has_damage_distribution = False
            elif row.startswith("SP 2+") and not current_section:
                units_data[current_unit_index]['SP 2+'] = strings[1]
                units_data[current_unit_index]['SP 2+ Description'] = ','.join(strings[2:])
                current_skill = "SP 2+"
                has_damage_distribution = False
            elif row.startswith("SP 2") and not current_section:
                units_data[current_unit_index]['SP 2'] = strings[1]
                units_data[current_unit_index]['SP 2 Description'] = ','.join(strings[2:])
                current_skill = "SP 2"
                has_damage_distribution = False
            elif row.startswith("SP 3") and not current_section:
                units_data[current_unit_index]['SP 3'] = strings[1]
                units_data[current_unit_index]['SP 3 Description'] = ','.join(strings[2:])
                current_skill = "SP 3"
                has_damage_distribution = False
            elif (row.startswith("SP") or row.startswith("Hidden Skill")) and not current_section:
                units_data[current_unit_index]['SP Skill'] = strings[1]
                units_data[current_unit_index]['SP Skill Description'] = ','.join(strings[2:])
                current_skill = "SP Skill"
                has_damage_distribution = False
            elif row.startswith("EX Skill") and not current_section:
                units_data[current_unit_index]['EX Skill'] = strings[1]
                units_data[current_unit_index]['EX Skill Description'] = ','.join(strings[2:])
                current_skill = None
                has_damage_distribution = False
            elif row.startswith('Unique Equipment 1') and not current_section:
                units_data[current_unit_index]['Unique Equipment 1'] = strings[1]
                stats = []
                for i in range(2, len(strings) - 1, 2):
                    stats.append(strings[i] + ' ' + strings[i+1])
                if stats:
                    units_data[current_unit_index]['Unique Equipment 1 Stats'] = stats
            elif row.startswith('Unique Equipment 2') and not current_section:
                units_data[current_unit_index]['Unique Equipment 2'] = strings[1]
                stats = []
                for i in range(2, len(strings) - 1, 2):
                    stats.append(strings[i] + ' ' + strings[i+1])
                if stats:
                    units_data[current_unit_index]['Unique Equipment 2 Stats'] = stats
            elif row.startswith('Misc. Information'):
                units_data[current_unit_index]['Misc. Information'] = ','.join(strings[1:])
                current_section = 'Misc. Information'
            elif row.startswith('Notes') or row.startswith('Pros'):
                units_data[current_unit_index]['Review'] = ','.join(strings[1:])
                current_section = 'Review'
            else:
                if has_damage_distribution:
                    units_data[current_unit_index][current_skill + ' Damage Distribution'] += '\n' + ' ' * (len('Damage Distribution') + 2) + row
                elif current_skill:
                    if units_data[current_unit_index][current_skill + ' Description'] and row:
                        units_data[current_unit_index][current_skill + ' Description'] += '\n' + row
                elif current_section and row:
                    units_data[current_unit_index][current_section] += '\n' + row
    
    with open('data.json', 'w') as file:
        json.dump(units_data, file, indent=4)
                
    return units_data, hashlib.md5(data).hexdigest()